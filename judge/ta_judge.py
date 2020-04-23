#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
MIT License

Copyright (c) 2020 Huang Po-Hsuan

Permission is hereby granted, free of charge, to any person obtaining a copy
of this software and associated documentation files (the "Software"), to deal
in the Software without restriction, including without limitation the rights
to use, copy, modify, merge, publish, distribute, sublicense, and/or sell
copies of the Software, and to permit persons to whom the Software is
furnished to do so, subject to the following conditions:

The above copyright notice and this permission notice shall be included in all
copies or substantial portions of the Software.

THE SOFTWARE IS PROVIDED "AS IS", WITHOUT WARRANTY OF ANY KIND, EXPRESS OR
IMPLIED, INCLUDING BUT NOT LIMITED TO THE WARRANTIES OF MERCHANTABILITY,
FITNESS FOR A PARTICULAR PURPOSE AND NONINFRINGEMENT. IN NO EVENT SHALL THE
AUTHORS OR COPYRIGHT HOLDERS BE LIABLE FOR ANY CLAIM, DAMAGES OR OTHER
LIABILITY, WHETHER IN AN ACTION OF CONTRACT, TORT OR OTHERWISE, ARISING FROM,
OUT OF OR IN CONNECTION WITH THE SOFTWARE OR THE USE OR OTHER DEALINGS IN THE
SOFTWARE.
"""

__version__ = '1.6.0'

from judge import ErrorHandler
from judge import LocalJudge
import judge
import argparse
import configparser
import os
from openpyxl import Workbook
from openpyxl import load_workbook
from zipfile import ZipFile
import rarfile
from glob import glob as globbing
import re
from collections import namedtuple
import logging
import sys
if sys.version_info < (3,):
    raise ImportError(
        "You are running local-judge {} on Python 2\n".format(__version__) +
        "Please use Python 3")


class TaJudge:
    def __init__(self, ta_config):
        try:
            self._config = ta_config['TaConfig']
            self.students_zip_container = self._config['StudentsZipContainer']
            self.students_pattern = self._config['StudentsPattern']
            self.students_extract_dir = self._config['StudentsExtractDir']
            self.extract_afresh = self._config['ExtractAfresh']
            self.students_zips = globbing(
                self.students_zip_container+os.sep+"*")

            # Parse the students' id and sort them
            self.students = self._parse_students()
            self.students.sort(key=lambda x: x.id)
        except KeyError as e:
            print("[ERROR] "+str(e) +
                  " field was not found in config file.")
            print("Please check `judge.conf` first.")
            exit(1)
        try:
            # Create the temporary directory for output
            # Suppress the error when the directory already exists
            os.makedirs(self.students_extract_dir)
        except OSError as e:
            if e.errno != os.errno.EEXIST:
                print("[ERROR] "+str(e))
                exit(1)

    def _parse_students(self):
        """ Used to parse student by the zip filename.

        The filename will be split into two parts: student id and the type of zip file (zip or rar).
        """
        Student = namedtuple(
            'Student', ('id', 'zip_type', 'zip_path', 'extract_path'))
        regex = re.compile(self.students_pattern)
        students = []
        for student_zip_path in self.students_zips:
            try:
                match = regex.search(student_zip_path)
                # group(1): whole filename for extracted path
                # group(2): student id
                # group(3): file type (zip or rar)
                students.append(Student(
                    match.group(2),
                    match.group(3),
                    os.path.abspath(student_zip_path),
                    os.path.abspath(self.students_extract_dir+os.sep+match.group(1))))
            except Exception as e:
                logging.error("Failed in parse stage. `" + str(student_zip_path) +
                              "` did not match the file rule.")
        return students

    def extract_student(self, student):
        """ Used to extract students' zip file.
        """
        if self.extract_afresh == "false":
            return
        try:
            if student.zip_type == 'zip':
                z = ZipFile(student.zip_path, 'r')
            elif student.zip_type == 'rar':
                z = rarfile.RarFile(student.zip_path)
            else:
                logging.error(str(
                    student.id)+" failed in extract stage with unknown zip type: `"+student.zip_type+"`")
                return
            z.extractall(self.students_extract_dir)
            z.close()
        except Exception as e:
            logging.error("`" + str(student.id) +
                          "` failed in extract stage. "+str(e))


def cd_student_path(student):
    """ Change directory to student source code.
    """
    try:
        os.chdir(student.extract_path)
    except FileNotFoundError as e:
        logging.error(str(student.id)+" "+str(e))


def find_student_by_id(student_id, students):
    """ Find the student from the students by id.
    """
    return [s for s in students if s.id == student_id]


def judge_one_student(tj, lj, student, here_path):
    """ Judge one student and return the correctness result.
    """
    judge.ERR_HANDLER.set_student_id(student.id)
    judge.ERR_HANDLER.clear_cache_log()
    tj.extract_student(student)
    cd_student_path(student)
    lj.build()
    correctness = []
    report_table = []
    correctness.append(student.id)
    for test in lj.tests:
        output = lj.run(test.input_filepath)
        accept, diff = lj.compare(output, test.answer_filepath)
        report_table.append(
            {'test': test.test_name, 'accept': accept, 'diff': diff})
        correctness.append(1 if accept else 0)
    # Restore the path
    os.chdir(here_path)
    return correctness, report_table


def append_log_msg(ori_result):
    if judge.ERR_HANDLER.cache_log != "":
        return ori_result + [1, judge.ERR_HANDLER.cache_log]
    else:
        return ori_result + [0]


def get_args():
    """ Init argparser and return the args from cli.
    """
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "-c", "--config",
        help="the config file",
        type=str,
        default=os.getcwd()+os.sep+"judge.conf")
    parser.add_argument(
        "-t", "--ta-config",
        help="the config file",
        type=str,
        default=os.getcwd()+os.sep+"ta_judge.conf")
    parser.add_argument(
        "-s", "--student",
        help="judge only one student",
        type=str,
        default=None)
    parser.add_argument(
        "-u", "--update",
        help="update specific student's score by rejudgement",
        type=str,
        default=None)
    return parser.parse_args()


if __name__ == '__main__':
    args = get_args()
    config = configparser.ConfigParser()
    ta_config = configparser.ConfigParser()
    config.read(args.config)
    ta_config.read(args.ta_config)

    logging_config = {
        'filename': 'ta_judge.log',
        'filemode': 'a',
        'format': '%(asctime)-15s [%(levelname)s] %(message)s'}
    judge.ERR_HANDLER = ErrorHandler('log', **logging_config)

    # Check if the config file is empty or not exist.
    if config.sections() == [] or ta_config.sections() == []:
        print("[ERROR] Failed in config stage.")
        raise FileNotFoundError(
            args.config if config.sections() == [] else args.ta_config)

    tj = TaJudge(ta_config)
    lj = LocalJudge(config)

    here_path = os.getcwd()
    students = tj.students


    if not args.student == None:
        # Assign specific student for this judgement and report to screen
        tj.extract_afresh = "false"
        students = find_student_by_id(args.student, tj.students)
        for student in students:
            result, report_table = judge_one_student(tj, lj, student, here_path)

        from judge import Report
        report = Report(report_verbose=True, total_score=config['Config']['TotalScore'])
        report.table = report_table
        report.print_report()
        exit(0)

    elif not args.update == None:
        # Update one student's judge result
        tj.extract_afresh = "false"
        students = find_student_by_id(args.update, tj.students)
        for student in students:
            result, report_table = judge_one_student(tj, lj, student, here_path)
        # Load existing table
        book = load_workbook(ta_config['TaConfig']['ScoreOutput'])
        sheet = book.active
        for row in sheet.rows:
            if row[0].value == args.update:
                for cell in row:
                    if sheet.cell(row=1, column=cell.column).value == "in_log":
                        continue
                    if cell.column-1 >= len(result):
                        break
                    cell.value = result[cell.column-1]
                break

    else:
        # Judge all students
        # Init the table with the title
        book = Workbook()
        sheet = book.active
        title = [t.test_name for t in lj.tests] + ['in_log', 'log_msg']
        title.insert(0, "student id")
        sheet.append(title)

        for student in students:
            result, report_table = judge_one_student(tj, lj, student, here_path)
            result = append_log_msg(result)
            sheet.append(result)

    book.save(ta_config['TaConfig']['ScoreOutput'])
    print("Finished")
